from __future__ import annotations

import io
import json
from datetime import datetime, timezone
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from cloud_backend import CloudDB
from config import MASTER_WORKBOOK_PATH
from ledger import append_transaction
from permissions import permission_summary, require

SHEET_TABLES = {
    "Source_Batches": "source_batch_view",
    "Tannery_Intake": "tannery_intakes",
    "AI_Inspections": "inspection_view",
    "Environmental": "environmental_records",
    "Processing": "processing_view",
    "Decisions": "decision_view",
    "Finished_Lots": "traceability_view",
    "Users_Access": "app_users",
    "Ledger": "ledger_transactions",
}


def _serialise(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def build_master_workbook_bytes(db: CloudDB) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, table in SHEET_TABLES.items():
        ws = wb.create_sheet(sheet_name)
        rows = db.select(table, order="id")
        if sheet_name == "Users_Access":
            for row in rows:
                row["permission_summary"] = permission_summary(str(row.get("role")))
                row.pop("password_hash", None); row.pop("password_salt", None)
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([_serialise(row.get(h)) for h in headers])
        else:
            ws.append(["No records yet"])
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="0B6B57")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        for col in range(1, ws.max_column + 1):
            longest = max(len(str(ws.cell(r, col).value or "")) for r in range(1, min(ws.max_row, 100) + 1))
            ws.column_dimensions[get_column_letter(col)].width = min(max(12, longest + 2), 38)
    out = io.BytesIO(); wb.save(out); return out.getvalue()


def sync_master_workbook(db: CloudDB, actor: dict[str, Any], reason: str) -> dict[str, Any]:
    data = build_master_workbook_bytes(db)
    db.upload_bytes(db.settings.system_bucket, MASTER_WORKBOOK_PATH, data,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", upsert=True)
    row = db.insert("excel_sync_log", {
        "sync_direction": "DATABASE_TO_EXCEL", "status": "SUCCESS", "reason": reason,
        "row_count": None, "actor_user_id": actor.get("id"),
        "actor_individual_id": actor.get("individual_user_id"),
        "synced_at": datetime.now(timezone.utc).isoformat(),
    })
    append_transaction(db, "MASTER_EXCEL_UPDATED", None, actor, {
        "path": MASTER_WORKBOOK_PATH, "reason": reason, "size_bytes": len(data)
    })
    return {"bytes": data, "path": MASTER_WORKBOOK_PATH, "log": row}


def import_source_workbook(db: CloudDB, actor: dict[str, Any], content: bytes) -> dict[str, Any]:
    require(actor, "excel.source_import")
    wb = load_workbook(io.BytesIO(content), data_only=True)
    if "Source_Batches" not in wb.sheetnames:
        raise ValueError("Workbook must contain a Source_Batches sheet")
    ws = wb["Source_Batches"]
    row1 = [str(c.value or "").strip() for c in ws[1]]
    row2 = [str(c.value or "").strip() for c in ws[2]] if ws.max_row >= 2 else []
    header_row = 1 if "source_type" in row1 else 2
    headers = row1 if header_row == 1 else row2
    required = {"source_type", "source_region", "collection_timestamp", "preservation_method",
                "salting_delay_hours", "hide_count"}
    missing = required - set(headers)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")
    imported = 0; errors: list[dict[str, Any]] = []
    from services import create_excel_batch_without_image
    for row_no, values in enumerate(ws.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
        if not any(v is not None and str(v).strip() for v in values):
            continue
        record = dict(zip(headers, values))
        try:
            create_excel_batch_without_image(db, actor, record, sync_excel=False)
            imported += 1
        except Exception as exc:
            errors.append({"row": row_no, "error": str(exc)})
    sync_master_workbook(db, actor, f"Excel import: {imported} source rows")
    db.insert("excel_sync_log", {
        "sync_direction": "EXCEL_TO_DATABASE", "status": "SUCCESS" if not errors else "PARTIAL",
        "reason": "Source_Batches bulk import", "row_count": imported,
        "actor_user_id": actor.get("id"), "actor_individual_id": actor.get("individual_user_id"),
        "error_report": errors, "synced_at": datetime.now(timezone.utc).isoformat(),
    })
    return {"imported": imported, "errors": errors}
